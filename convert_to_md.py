import os
import sys
import traceback

if sys.platform == "win32":
    import ctypes

    ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    ctypes.windll.kernel32.SetConsoleCP(65001)
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

SUPPORTED_PDF = {".pdf"}
SUPPORTED_EXCEL = {".xlsx", ".xlsm", ".xls"}


def log(msg):
    print(msg, flush=True)


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip().replace("|", "\\|").replace("\n", "<br>")


def rows_to_markdown_table(rows):
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return "_(빈 시트)_"
    header = [cell_text(c) for c in rows[0]]
    width = len(header)
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in rows[1:]:
        row = list(row) + [None] * (width - len(row))
        lines.append("| " + " | ".join(cell_text(c) for c in row[:width]) + " |")
    return "\n".join(lines)


def table_to_markdown(table):
    rows = [["" if c is None else str(c).strip() for c in row] for row in table]
    return rows_to_markdown_table(rows)


def is_blank(value):
    return value is None or str(value).strip() == ""


def is_table_reliable(table):
    if len(table) < 2:
        return False
    cells = [c for row in table for c in row]
    if not cells:
        return False

    empty = sum(1 for c in cells if is_blank(c))
    if (empty / len(cells)) > 0.25:
        return False

    header, data_rows = table[0], table[1:]
    for col in range(len(header)):
        values = [row[col] for row in data_rows if col < len(row)]
        if values and all(is_blank(v) for v in values):
            return False

    return True


def pdf_to_markdown(path):
    import pdfplumber

    parts = []
    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            parts.append(f"## Page {page_num}\n")

            for region in split_columns(page):
                text = (region.extract_text() or "").strip()
                if text:
                    parts.append(text)

                table_count = 0
                for table in region.extract_tables():
                    if not is_table_reliable(table):
                        continue
                    md_table = table_to_markdown(table)
                    if md_table:
                        table_count += 1
                        parts.append(f"**표 {table_count}**\n\n{md_table}")

    return "\n\n".join(parts).strip() + "\n"


def detect_column_split(page):
    words = page.extract_words()
    if not words:
        return None

    width = page.width
    lo, hi = width * 0.25, width * 0.75
    intervals = sorted(
        (w["x0"], w["x1"]) for w in words if w["x1"] > lo and w["x0"] < hi
    )

    gaps = []
    prev_end = lo
    for x0, x1 in intervals:
        if x0 > prev_end:
            gaps.append((prev_end, x0))
        prev_end = max(prev_end, x1)
    if hi > prev_end:
        gaps.append((prev_end, hi))

    min_gap = width * 0.03
    candidates = [g for g in gaps if g[1] - g[0] >= min_gap]
    if not candidates:
        return None

    best = max(candidates, key=lambda g: g[1] - g[0])
    return (best[0] + best[1]) / 2


def split_columns(page):
    split_x = detect_column_split(page)
    if split_x is None:
        return [page]
    return [
        page.crop((0, 0, split_x, page.height)),
        page.crop((split_x, 0, page.width, page.height)),
    ]


def excel_to_markdown(path):
    ext = os.path.splitext(path)[1].lower()
    sections = []

    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            sections.append(f"## {sheet.title}\n\n{rows_to_markdown_table(rows)}")
    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        for sheet in wb.sheets():
            rows = [sheet.row_values(r) for r in range(sheet.nrows)]
            sections.append(f"## {sheet.name}\n\n{rows_to_markdown_table(rows)}")
    else:
        raise ValueError(f"지원하지 않는 엑셀 확장자: {ext}")

    return "\n\n".join(sections).strip() + "\n"


def find_target_folders():
    folders = []
    for arg in sys.argv[1:]:
        if os.path.isdir(arg):
            folders.append(arg)
        elif os.path.isfile(arg):
            folders.append(os.path.dirname(arg))

    if not folders:
        if getattr(sys, "frozen", False):
            folders.append(os.path.dirname(sys.executable))
        else:
            folders.append(os.path.dirname(os.path.abspath(__file__)))

    seen = set()
    unique = []
    for f in folders:
        if f not in seen:
            seen.add(f)
            unique.append(f)
    return unique


def process_folder(folder):
    log(f"\n대상 폴더: {folder}")
    log("=" * 60)

    converted = 0
    failed = 0

    for root, _, files in os.walk(folder):
        for name in files:
            ext = os.path.splitext(name)[1].lower()
            if ext not in SUPPORTED_PDF and ext not in SUPPORTED_EXCEL:
                continue

            src = os.path.join(root, name)
            md_path = src + ".md"

            try:
                if ext in SUPPORTED_PDF:
                    content = pdf_to_markdown(src)
                else:
                    content = excel_to_markdown(src)

                with open(md_path, "w", encoding="utf-8") as f:
                    f.write(content)

                log(f"[성공] {src} -> {md_path}")
                converted += 1
            except Exception as e:
                log(f"[실패] {src}: {e}")
                traceback.print_exc()
                failed += 1

    log("=" * 60)
    log(f"폴더 완료: 성공 {converted}개, 실패 {failed}개")
    return converted, failed


def main():
    folders = find_target_folders()

    total_converted = 0
    total_failed = 0
    for folder in folders:
        c, f = process_folder(folder)
        total_converted += c
        total_failed += f

    log("\n" + "=" * 60)
    log(f"전체 완료: 성공 {total_converted}개, 실패 {total_failed}개")


if __name__ == "__main__":
    main()
    try:
        input("\n엔터 키를 누르면 종료합니다...")
    except EOFError:
        pass
